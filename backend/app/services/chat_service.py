import httpx
import io
from loguru import logger
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.encryption import decrypt_api_key
from app.db.models import Bot, Message
from app.repositories import bot as bot_repo
from app.schemas.chat import ChatResponse

from pdfminer.high_level import extract_text as pdf_extract
from openpyxl import load_workbook
from docx import Document


async def extract_text_from_file(filename: str, content: bytes) -> str:
    """Extract text from PDF, XLS, DOC, TXT files."""
    ext = filename.lower().split('.')[-1] if '.' in filename else ''

    try:
        if ext == 'pdf':
            return pdf_extract(io.BytesIO(content))
        elif ext in ['xls', 'xlsx']:
            wb = load_workbook(filename=io.BytesIO(content), read_only=True)
            text = []
            for sheet in wb:
                for row in sheet.iter_rows(values_only=True):
                    text.append(' '.join(str(c) for c in row if c is not None))
            return '\n'.join(text)
        elif ext == 'docx':
            doc = Document(io.BytesIO(content))
            return '\n'.join(p.text for p in doc.paragraphs)
        elif ext == 'txt':
            return content.decode('utf-8', errors='ignore')
        else:
            return f"[Unsupported file type: {ext}]"
    except Exception as e:
        logger.error(f"File extraction error: {e}")
        return f"[Error extracting file: {filename}]"


async def process_chat(db: AsyncSession, message) -> ChatResponse:
    bot = await bot_repo.get_by_id(db, message.bot_id)
    if not bot or not bot.is_active:
        raise ValueError("Bot not found or inactive")

    thread_id = message.thread_id or "default"

    # Extract and append file contents if present
    content = message.query
    if message.files:
        file_texts = []
        for i, f in enumerate(message.files):
            # Try to get filename from somewhere, or use generic name
            filename = f"file_{i}"
            text = await extract_text_from_file(filename, f)
            file_texts.append(f"--- File {filename} ---\n{text}")
        content += "\n\n" + "\n\n".join(file_texts)

    await save_message(db, bot.id, thread_id, "user", content)

    try:
        answer = await call_llm(bot, content, thread_id)
        answer += (
            "\n\n*Справка носит информационный характер и не заменяет официальную консультацию.*"
        )

        await save_message(db, bot.id, thread_id, "assistant", answer)

        return ChatResponse(answer=answer, thread_id=thread_id, sources=[])
    except Exception as e:
        logger.error(f"Chat error: {e}")
        raise


async def call_llm(bot: Bot, query: str, thread_id: str) -> str:
    api_key = decrypt_api_key(bot.api_key_encrypted) if bot.api_key_encrypted else ""

    if not api_key:
        return "Ошибка: API-ключ не настроен. Откройте настройки бота и добавьте ключ."

    if bot.provider == "openai":
        return await call_openai(bot.model, api_key, bot.system_prompt, query)
    elif bot.provider == "anthropic":
        return await call_anthropic(bot.model, api_key, bot.system_prompt, query)
    else:
        return await call_openai_compatible(
            bot.provider, bot.model, api_key, bot.system_prompt, query
        )


async def call_openai(model: str, api_key: str, system: str, query: str) -> str:
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    payload = {
        "model": model,
        "messages": [{"role": "system", "content": system}, {"role": "user", "content": query}],
        "stream": False,
    }

    async with httpx.AsyncClient(timeout=60.0) as client:
        resp = await client.post(
            "https://api.openai.com/v1/chat/completions", json=payload, headers=headers
        )
        resp.raise_for_status()
        data = resp.json()
        return data["choices"][0]["message"]["content"]


async def call_anthropic(model: str, api_key: str, system: str, query: str) -> str:
    headers = {
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
        "Content-Type": "application/json",
    }
    payload = {
        "model": model,
        "system": system,
        "messages": [{"role": "user", "content": query}],
        "max_tokens": 4096,
    }

    async with httpx.AsyncClient(timeout=60.0) as client:
        resp = await client.post(
            "https://api.anthropic.com/v1/messages", json=payload, headers=headers
        )
        resp.raise_for_status()
        data = resp.json()
        return data["content"][0]["text"]


async def call_openai_compatible(
    provider: str, model: str, api_key: str, system: str, query: str
) -> str:
    base_urls = {
        "groq": "https://api.groq.com/openai/v1",
        "openrouter": "https://openrouter.ai/api/v1",
        "deepseek": "https://api.deepseek.com/v1",
    }
    base_url = base_urls.get(provider, provider)

    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    payload = {
        "model": model,
        "messages": [{"role": "system", "content": system}, {"role": "user", "content": query}],
    }

    async with httpx.AsyncClient(timeout=60.0) as client:
        resp = await client.post(f"{base_url}/chat/completions", json=payload, headers=headers)
        resp.raise_for_status()
        data = resp.json()
        return data["choices"][0]["message"]["content"]


async def save_message(db: AsyncSession, bot_id: int, thread_id: str, role: str, content: str):
    msg = Message(bot_id=bot_id, thread_id=thread_id, role=role, content=content)
    db.add(msg)
    await db.commit()


async def test_llm_connection(bot: Bot) -> tuple[bool, str]:
    """Test LLM connectivity. Returns (success, message)."""
    api_key = decrypt_api_key(bot.api_key_encrypted) if bot.api_key_encrypted else ""
    if not api_key:
        return False, "API-ключ не настроен"

    try:
        if bot.provider == "openai":
            return await test_openai(bot.model, api_key, bot.system_prompt)
        elif bot.provider == "anthropic":
            return await test_anthropic(bot.model, api_key, bot.system_prompt)
        else:
            return await test_openai_compatible(bot.provider, bot.model, api_key, bot.system_prompt)
    except Exception as e:
        return False, f"Ошибка подключения: {e}"


async def test_openai(model: str, api_key: str, system: str) -> tuple[bool, str]:
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    payload = {
        "model": model,
        "messages": [{"role": "system", "content": system}, {"role": "user", "content": "test"}],
        "max_tokens": 1,
    }
    async with httpx.AsyncClient(timeout=10.0) as client:
        resp = await client.post("https://api.openai.com/v1/chat/completions", json=payload, headers=headers)
        resp.raise_for_status()
    return True, "OK"


async def test_anthropic(model: str, api_key: str, system: str) -> tuple[bool, str]:
    headers = {"x-api-key": api_key, "anthropic-version": "2023-06-01", "Content-Type": "application/json"}
    payload = {"model": model, "system": system, "messages": [{"role": "user", "content": "test"}], "max_tokens": 1}
    async with httpx.AsyncClient(timeout=10.0) as client:
        resp = await client.post("https://api.anthropic.com/v1/messages", json=payload, headers=headers)
        resp.raise_for_status()
    return True, "OK"


async def test_openai_compatible(provider: str, model: str, api_key: str, system: str) -> tuple[bool, str]:
    base_urls = {
        "groq": "https://api.groq.com/openai/v1",
        "openrouter": "https://openrouter.ai/api/v1",
        "deepseek": "https://api.deepseek.com/v1",
    }
    base_url = base_urls.get(provider, provider)
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    payload = {
        "model": model,
        "messages": [{"role": "system", "content": system}, {"role": "user", "content": "test"}],
        "max_tokens": 1,
    }
    async with httpx.AsyncClient(timeout=10.0) as client:
        resp = await client.post(f"{base_url}/chat/completions", json=payload, headers=headers)
        resp.raise_for_status()
    return True, "OK"
